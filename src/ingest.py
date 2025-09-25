import pandas as pd
import numpy as np
import streamlit as st
from pathlib import Path
import openpyxl
import re
from typing import Union, Dict, List

def detect_largest_sheet(file_path: Union[str, Path]) -> str:
    """
    Detect the sheet with the most data in an Excel file.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Name of the sheet with most data
    """
    try:
        # Load workbook to get sheet names and sizes
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_sizes = {}
        
        for sheet_name in wb.sheetnames:
            try:
                # Quick read to get dimensions
                df_temp = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                full_df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_sizes[sheet_name] = len(full_df) * len(full_df.columns)
            except Exception as e:
                print(f"Error reading sheet {sheet_name}: {e}")
                sheet_sizes[sheet_name] = 0
        
        wb.close()
        
        # Return sheet with maximum data
        if sheet_sizes:
            return max(sheet_sizes.keys(), key=lambda k: sheet_sizes[k])
        else:
            return wb.sheetnames[0] if wb.sheetnames else None
            
    except Exception as e:
        print(f"Error detecting largest sheet: {e}")
        return None

def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize column names to standard format.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with normalized column names
    """
    df = df.copy()
    
    # Column mapping dictionary
    column_mapping = {
        # Entity identification
        'entity': 'entity_name',
        'entity_name': 'entity_name',
        'state_entity': 'entity_name',  # Fix for our dataset
        'organisation': 'entity_name',
        'organization': 'entity_name',
        'institution': 'entity_name',
        'company': 'entity_name',
        'name': 'entity_name',
        
        # Year
        'year': 'year',
        'financial_year': 'year',
        'reporting_year': 'year',
        'period': 'year',
        
        # Financial metrics
        'revenue': 'revenue',
        'total_revenue': 'revenue',
        'income': 'revenue',
        'total_income': 'revenue',
        
        'expenditure': 'expenditure',
        'expenses': 'expenditure',
        'total_expenditure': 'expenditure',
        'total_expenses': 'expenditure',
        'spending': 'expenditure',
        
        'assets': 'total_assets',
        'total_assets': 'total_assets',
        
        'liabilities': 'total_liabilities',
        'total_liabilities': 'total_liabilities',
        
        'equity': 'equity',
        'net_equity': 'equity',
        'shareholders_equity': 'equity',
        
        # Cash and liquidity
        'cash': 'cash',
        'cash_and_equivalents': 'cash',
        'liquid_assets': 'cash',
        
        # Procurement
        'contracts': 'contract_count',
        'contract_value': 'contract_value',
        'procurement_value': 'contract_value',
        'single_source': 'single_source_contracts',
        
        # Employee related
        'employees': 'employee_count',
        'staff': 'employee_count',
        'payroll': 'payroll_expense',
        'salaries': 'payroll_expense',
        
        # Regional
        'region': 'region',
        'location': 'region',
        'district': 'region',
        
        # Type/Category
        'type': 'entity_type',
        'category': 'entity_type',
        'sector': 'sector',
        'ministry': 'ministry',
        
        # Target variable (audit flag)
        'audit_flag': 'audit_flag',
        'dependet_variable': 'audit_flag',  # Fix typo in dataset
        'dependent_variable': 'audit_flag',
        'target': 'audit_flag',
        'irregularity_flag': 'audit_flag',
        'flag': 'audit_flag',
        
        # Dataset specific columns
        'operational_revenue': 'operational_revenue',
        'other_income': 'other_income',
        'total_revenue': 'total_revenue',
        'direct_costs': 'direct_costs',
        'admin_exp': 'admin_expense',
        'finance_cost': 'finance_cost',
        'total_expenditure': 'total_expenditure',
        'gross_profit_losses': 'gross_profit',
        'net_profit_losses': 'net_profit',
        'non_current_assets': 'non_current_assets',
        'current_assets': 'current_assets',
        'current_liabilities': 'current_liabilities',
        'net_assets': 'net_assets'
    }
    
    # Normalize column names (lowercase, remove special chars)
    normalized_cols = {}
    for col in df.columns:
        # Convert to lowercase and remove special characters
        clean_col = re.sub(r'[^a-zA-Z0-9_]', '', str(col).lower().strip())
        clean_col = re.sub(r'_+', '_', clean_col)  # Remove multiple underscores
        clean_col = clean_col.strip('_')  # Remove leading/trailing underscores
        
        # Map to standard name if available
        if clean_col in column_mapping:
            normalized_cols[col] = column_mapping[clean_col]
        else:
            normalized_cols[col] = clean_col
    
    # Rename columns
    df = df.rename(columns=normalized_cols)
    
    return df

def clean_data_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and convert data types appropriately.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with cleaned data types
    """
    df = df.copy()
    
    # Numeric columns that should be converted
    numeric_columns = [
        'revenue', 'expenditure', 'total_assets', 'total_liabilities', 
        'equity', 'cash', 'contract_value', 'payroll_expense', 
        'employee_count', 'contract_count', 'single_source_contracts'
    ]
    
    for col in numeric_columns:
        if col in df.columns:
            # Remove currency symbols and commas
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).str.replace(r'[₵$,\s]', '', regex=True)
                df[col] = df[col].str.replace(r'[^\d.-]', '', regex=True)
            
            # Convert to numeric
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Year column
    if 'year' in df.columns:
        df['year'] = pd.to_numeric(df['year'], errors='coerce')
        # Filter reasonable years
        df = df[(df['year'] >= 2000) & (df['year'] <= 2030)]
    
    # Entity name cleaning
    if 'entity_name' in df.columns:
        df['entity_name'] = df['entity_name'].astype(str).str.strip()
        df['entity_name'] = df['entity_name'].str.title()  # Proper case
        # Remove entities with invalid names
        df = df[df['entity_name'].str.len() > 2]
        df = df[~df['entity_name'].isin(['Nan', 'None', 'null', ''])]
    
    return df

def load_dataset(file_path: Union[str, Path, object]) -> pd.DataFrame:
    """
    Load dataset from Excel or CSV file with automatic preprocessing.
    
    Args:
        file_path: Path to file or uploaded file object
        
    Returns:
        Processed DataFrame
    """
    try:
        # Handle Streamlit uploaded file
        if hasattr(file_path, 'read'):
            if file_path.name.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                # For Excel files, try to detect the best sheet
                df = pd.read_excel(file_path)
        else:
            # Handle file path
            file_path = Path(file_path)
            
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if file_path.suffix.lower() == '.csv':
                df = pd.read_csv(file_path)
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                # Detect largest sheet
                sheet_name = detect_largest_sheet(file_path)
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    df = pd.read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        print(f"Loaded dataset with shape: {df.shape}")
        
        # Normalize column names
        df = normalize_column_names(df)
        
        # Clean data types
        df = clean_data_types(df)
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Basic validation
        if len(df) == 0:
            raise ValueError("Dataset is empty after cleaning")
        
        if 'entity_name' not in df.columns:
            raise ValueError("No entity name column found. Please ensure your dataset has entity/organization names.")
        
        print(f"Processed dataset shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Save processed data
        output_dir = Path("data/processed")
        output_dir.mkdir(parents=True, exist_ok=True)
        df.to_parquet(output_dir / "model_dataset.parquet", index=False)
        
        return df
        
    except Exception as e:
        print(f"Error loading dataset: {e}")
        raise e

def get_dataset_info(df: pd.DataFrame) -> Dict:
    """
    Get comprehensive information about the dataset.
    
    Args:
        df: Input DataFrame
        
    Returns:
        Dictionary with dataset information
    """
    info = {
        'shape': df.shape,
        'columns': list(df.columns),
        'numeric_columns': list(df.select_dtypes(include=[np.number]).columns),
        'categorical_columns': list(df.select_dtypes(include=['object']).columns),
        'missing_values': df.isnull().sum().to_dict(),
        'memory_usage': df.memory_usage(deep=True).sum(),
    }
    
    # Year range
    if 'year' in df.columns:
        info['year_range'] = (df['year'].min(), df['year'].max())
    
    # Entity count
    if 'entity_name' in df.columns:
        info['unique_entities'] = df['entity_name'].nunique()
    
    return info
